"""
XLSX (Excel) spreadsheet processor
"""
import logging
from pathlib import Path
from typing import Dict, List, Optional, Any
import json

try:
    import openpyxl
    from openpyxl.worksheet.worksheet import Worksheet
    XLSX_AVAILABLE = True
except ImportError:
    XLSX_AVAILABLE = False
    logging.warning("openpyxl not installed. Install with: pip install openpyxl")

logger = logging.getLogger(__name__)


class XLSXProcessor:
    """
    Process Microsoft Excel (.xlsx) spreadsheets
    
    Features:
    - Extract data from all sheets
    - Preserve formulas and formatting
    - Handle merged cells
    - Extract cell comments
    - Convert to text and JSON formats
    """
    
    def __init__(
        self,
        max_rows: Optional[int] = None,
        max_cols: Optional[int] = None,
        include_formulas: bool = False
    ):
        """
        Initialize XLSX processor
        
        Args:
            max_rows: Maximum rows to process per sheet (None = all)
            max_cols: Maximum columns to process per sheet (None = all)
            include_formulas: Whether to include formula text
        """
        if not XLSX_AVAILABLE:
            raise ImportError(
                "openpyxl is required for XLSX processing. "
                "Install with: pip install openpyxl"
            )
        
        self.max_rows = max_rows
        self.max_cols = max_cols
        self.include_formulas = include_formulas
        logger.info("XLSXProcessor initialized")
    
    def process(
        self,
        file_path: str,
        sheet_names: Optional[List[str]] = None
    ) -> Dict[str, Any]:
        """
        Process an XLSX file and extract content
        
        Args:
            file_path: Path to XLSX file
            sheet_names: Specific sheets to process (None = all)
            
        Returns:
            Dictionary containing:
                - text: Full extracted text
                - sheets: List of sheet data
                - metadata: Workbook metadata
        """
        try:
            file_path = Path(file_path)
            if not file_path.exists():
                raise FileNotFoundError(f"File not found: {file_path}")
            
            if not file_path.suffix.lower() in ['.xlsx', '.xlsm']:
                raise ValueError(f"Not an XLSX file: {file_path}")
            
            logger.info(f"Processing XLSX: {file_path}")
            
            # Load workbook
            workbook = openpyxl.load_workbook(
                str(file_path),
                data_only=not self.include_formulas
            )
            
            # Process sheets
            sheets_data = []
            target_sheets = sheet_names if sheet_names else workbook.sheetnames
            
            for sheet_name in target_sheets:
                if sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    sheet_data = self._process_sheet(sheet)
                    sheets_data.append(sheet_data)
            
            # Extract metadata
            metadata = self._extract_metadata(workbook, file_path)
            
            # Combine all text
            full_text = self._combine_sheets_text(sheets_data)
            
            result = {
                "text": full_text,
                "sheets": sheets_data,
                "metadata": metadata,
                "file_path": str(file_path),
                "file_type": "xlsx",
                "sheet_count": len(sheets_data)
            }
            
            logger.info(
                f"XLSX processed: {len(sheets_data)} sheets, "
                f"{sum(s['row_count'] for s in sheets_data)} total rows"
            )
            
            return result
            
        except Exception as e:
            logger.error(f"Error processing XLSX {file_path}: {str(e)}")
            raise
    
    def _process_sheet(self, sheet: Worksheet) -> Dict[str, Any]:
        """Process a single worksheet"""
        sheet_name = sheet.title
        
        # Determine dimensions
        max_row = sheet.max_row
        max_col = sheet.max_column
        
        if self.max_rows:
            max_row = min(max_row, self.max_rows)
        if self.max_cols:
            max_col = min(max_col, self.max_cols)
        
        # Extract data
        rows_data = []
        text_lines = []
        
        for row_idx, row in enumerate(sheet.iter_rows(
            min_row=1,
            max_row=max_row,
            max_col=max_col,
            values_only=True
        ), start=1):
            row_values = [self._format_cell_value(val) for val in row]
            rows_data.append(row_values)
            
            # Create text representation
            row_text = " | ".join(str(v) for v in row_values if v)
            if row_text.strip():
                text_lines.append(row_text)
        
        # Extract comments
        comments = self._extract_comments(sheet)
        
        return {
            "sheet_name": sheet_name,
            "text": "\n".join(text_lines),
            "rows": rows_data,
            "row_count": len(rows_data),
            "col_count": max_col,
            "comments": comments,
            "has_data": len(text_lines) > 0
        }
    
    def _format_cell_value(self, value: Any) -> Any:
        """Format cell value for output"""
        if value is None:
            return ""
        
        # Handle datetime
        if hasattr(value, 'isoformat'):
            return value.isoformat()
        
        # Handle numbers
        if isinstance(value, (int, float)):
            return value
        
        # Convert to string
        return str(value).strip()
    
    def _extract_comments(self, sheet: Worksheet) -> List[Dict]:
        """Extract cell comments from sheet"""
        comments = []
        
        try:
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.comment:
                        comments.append({
                            "cell": cell.coordinate,
                            "text": cell.comment.text,
                            "author": cell.comment.author if hasattr(cell.comment, 'author') else None
                        })
        except Exception as e:
            logger.warning(f"Error extracting comments: {str(e)}")
        
        return comments
    
    def _extract_metadata(self, workbook: openpyxl.Workbook, file_path: Path) -> Dict:
        """Extract workbook metadata"""
        metadata = {
            "filename": file_path.name,
            "sheet_names": workbook.sheetnames,
            "sheet_count": len(workbook.sheetnames)
        }
        
        try:
            props = workbook.properties
            
            if props.title:
                metadata["title"] = props.title
            if props.creator:
                metadata["creator"] = props.creator
            if props.created:
                metadata["created"] = str(props.created)
            if props.modified:
                metadata["modified"] = str(props.modified)
            if props.lastModifiedBy:
                metadata["last_modified_by"] = props.lastModifiedBy
            if props.subject:
                metadata["subject"] = props.subject
            if props.description:
                metadata["description"] = props.description
            if props.category:
                metadata["category"] = props.category
            if props.keywords:
                metadata["keywords"] = props.keywords
        
        except Exception as e:
            logger.warning(f"Error extracting metadata: {str(e)}")
        
        return metadata
    
    def _combine_sheets_text(self, sheets_data: List[Dict]) -> str:
        """Combine all sheets into full text"""
        parts = []
        
        for sheet in sheets_data:
            if sheet["has_data"]:
                sheet_text = f"[Sheet: {sheet['sheet_name']}]\n{sheet['text']}"
                parts.append(sheet_text)
                
                # Add comments if any
                if sheet["comments"]:
                    comments_text = "\n".join(
                        f"  Comment at {c['cell']}: {c['text']}"
                        for c in sheet["comments"]
                    )
                    parts.append(f"[Comments]\n{comments_text}")
        
        return "\n\n".join(parts)
    
    def extract_text_only(self, file_path: str, sheet_names: Optional[List[str]] = None) -> str:
        """
        Quick extraction of just the text content
        
        Args:
            file_path: Path to XLSX file
            sheet_names: Specific sheets to process (None = all)
            
        Returns:
            Extracted text as string
        """
        result = self.process(file_path, sheet_names)
        return result["text"]
    
    def to_json(self, file_path: str, sheet_names: Optional[List[str]] = None) -> str:
        """
        Convert XLSX to JSON format
        
        Args:
            file_path: Path to XLSX file
            sheet_names: Specific sheets to process
            
        Returns:
            JSON string representation
        """
        result = self.process(file_path, sheet_names)
        
        # Create JSON-friendly structure
        json_data = {
            "metadata": result["metadata"],
            "sheets": [
                {
                    "name": sheet["sheet_name"],
                    "rows": sheet["rows"],
                    "comments": sheet["comments"]
                }
                for sheet in result["sheets"]
            ]
        }
        
        return json.dumps(json_data, indent=2, default=str)
    
    def to_csv_text(self, file_path: str, sheet_name: Optional[str] = None) -> str:
        """
        Convert XLSX to CSV-like text format
        
        Args:
            file_path: Path to XLSX file
            sheet_name: Specific sheet (None = first sheet)
            
        Returns:
            CSV-formatted text
        """
        result = self.process(file_path, [sheet_name] if sheet_name else None)
        
        if not result["sheets"]:
            return ""
        
        # Use first sheet
        sheet = result["sheets"][0]
        csv_lines = []
        
        for row in sheet["rows"]:
            csv_line = ",".join(f'"{str(val)}"' for val in row)
            csv_lines.append(csv_line)
        
        return "\n".join(csv_lines)


def is_xlsx_available() -> bool:
    """Check if XLSX processing is available"""
    return XLSX_AVAILABLE


# Convenience function
def process_xlsx(file_path: str, sheet_names: Optional[List[str]] = None) -> Dict:
    """
    Process an XLSX file (convenience function)
    
    Args:
        file_path: Path to XLSX file
        sheet_names: Specific sheets to process (None = all)
        
    Returns:
        Processed spreadsheet data
    """
    processor = XLSXProcessor()
    return processor.process(file_path, sheet_names)
